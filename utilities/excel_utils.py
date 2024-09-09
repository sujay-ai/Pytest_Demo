import openpyxl

def get_row_count(file_path, sheet_name):
    """Returns the number of rows in the specified Excel sheet."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(file_path, sheet_name):
    """Returns the number of columns in the specified Excel sheet."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column

def read_data(file_path, sheet_name, row_num, col_num):
    """Reads data from a specified cell in the Excel sheet."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=col_num).value

def write_data(file_path, sheet_name, row_num, col_num, data):
    """Writes data to a specified cell in the Excel sheet."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(file_path)
